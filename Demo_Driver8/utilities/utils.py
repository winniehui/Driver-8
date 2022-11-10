import os
from time import sleep

import softest
from openpyxl import Workbook,load_workbook

current_path = os.path.abspath('.')
data_path = current_path + "\\testdata\\"
result_path = current_path + "\\reports\\"

class Utils(softest.TestCase):
    def get_value(self):
        with open(data_path + "D8.txt", "r", errors="ignore") as f:
            lines = f.readlines()
            p_dark = "file[" + lines[11].split("\n")[0] + "]"
            p_speed = "file[" + lines[12].split("\n")[0] + "]"
            p_width = "file[" + lines[20].split("\n")[0] + "]"
            p_height = "file[" + lines[21].split("\n")[0] + "]"
            print("p_dark: ", p_dark)
            print("p_speed: ", p_speed)
            print("p_width: ", p_width)
            print("p_height: ", p_height)
            f.close()
            return p_dark,p_speed,p_width,p_height

    def get_wh(self, pre_default_width, pre_default_height, t_width, t_height):
        v_width = ''
        v_height = ''

        if int(pre_default_width) == int(t_width):
            v_width = "E[" + str(pre_default_width) + "]"
        else:
            v_width = "NE[expected value:" + str(pre_default_width) + ", actual value:" + str(t_width) + "]"

        if int(pre_default_height) == int(t_height):
            v_height = "E[" + str(pre_default_height) + "]"
        else:
            v_height = "NE[expected value:" + str(pre_default_height) + ", actual value:" + str(t_height) + "]"

        # 获取save to file的值，并写入结果表里面
        f_dark, f_speed, f_width, f_height = self.get_value()
        v_width = v_width + "\n" + f_width
        v_height = v_height + "\n" + f_height
        return v_width, v_height

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(3, 5):
        #for i in range(3, col_ct + 1):
            col = []
            for j in range(5, 16):
                col.append(sh.cell(row=j, column=i).value)
            col.append(i)
            datalist.append(col)
        return datalist

    def write_data_to_excel_wh(self, v_width, v_height, i, j):
        result_name = result_path + "Report8_SRSmapping_Forerunner&Stingray.xlsx"
        # 写入指定文件
        wb = load_workbook(filename=result_name)  ##读取路径
        ws = wb.active
        # 结果值写入表格
        ws.cell(row=i, column=j, value=v_width)
        ws.cell(row=i+1, column=j, value=v_height)
        sleep(2)

        wb.save(result_name)

    def get_sd(self, pre_default_speed, pre_SpeedLimits, pre_default_darkness, pre_DarknessLimits,
               t_speed, t_SpeedLimits, t_darkness, t_DarknessLimits,p_speed,p_dark):
        # 结果判断
        v_speed = ''
        v_SpeedLimits = ''
        v_darkness = ''
        v_DarknessLimits = ''

        if int(pre_default_speed) == int(t_speed):
            v_speed = "E[" + str(pre_default_speed) + "]"
        else:
            v_speed = "NE[expected value:" + str(pre_default_speed) + ", actual value:" + str(t_speed) + "]"

        if pre_SpeedLimits == t_SpeedLimits:
            v_SpeedLimits = "E[" + pre_SpeedLimits + "]"
        else:
            v_SpeedLimits = "NE[expected value:" + pre_SpeedLimits + ", actual value:" + t_SpeedLimits + "]"

        if int(pre_default_darkness) == int(t_darkness):
            v_darkness = "E[" + str(pre_default_darkness) + "]"
        else:
            v_darkness = "NE[expected value:" + str(pre_default_darkness) + ", actual value:" + str(t_darkness) + "]"

        if pre_DarknessLimits == t_DarknessLimits:
            v_DarknessLimits = "E[" + pre_DarknessLimits + "]"
        else:
            v_DarknessLimits = "NE[expected value:" + pre_DarknessLimits + ", actual value:" + t_DarknessLimits + "]"

        v_speed = v_speed + "\n" + p_speed
        v_darkness = v_darkness + "\n" + p_dark
        return v_speed, v_darkness, v_SpeedLimits, v_DarknessLimits

    def write_data_to_excel_sd(self, v_speed, v_darkness, v_SpeedLimits, v_DarknessLimits, j):
        result_name = result_path + "Report8_SRSmapping_Forerunner&Stingray.xlsx"
        # 写入指定文件
        wb = load_workbook(filename=result_name)  ##读取路径
        ws = wb.active
        # 结果值写入表格
        ws.cell(row=8, column=j, value=v_speed)
        ws.cell(row=14, column=j, value=v_SpeedLimits)
        ws.cell(row=9, column=j, value=v_darkness)
        ws.cell(row=15, column=j, value=v_DarknessLimits)
        sleep(2)

        wb.save(result_name)
